#!/usr/bin/env python3
"""
DynamoDB에 Q&A 데이터 임베딩 스크립트

용도:
1. Q&A.xlsx 파일 읽기
2. 각 질문을 Bedrock Titan Embeddings로 변환
3. DynamoDB에 저장

실행:
python scripts/ingest_dynamodb.py data/Q&A.xlsx
"""

import os
import sys
import json
import logging
import uuid
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("⚠️  openpyxl 필요. 설치 중...")
    os.system("pip install openpyxl")
    import openpyxl

try:
    import boto3
except ImportError:
    print("⚠️  boto3 필요. 설치 중...")
    os.system("pip install boto3")
    import boto3

from dotenv import load_dotenv

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# 환경 변수 로드
load_dotenv()

# 설정
BEDROCK_REGION = os.environ.get("BEDROCK_REGION", "ap-northeast-1")
BEDROCK_MODEL_ID = os.environ.get("BEDROCK_MODEL_ID", "amazon.titan-embed-text-v1")
DYNAMODB_TABLE = os.environ.get("DYNAMODB_TABLE", "qa-documents")
EXCEL_FILE = "data/Q&A.xlsx"
SHEET_NAME = 0  # 첫 번째 시트


class QAIngestor:
    def __init__(self):
        """클라이언트 초기화"""
        self.bedrock = boto3.client("bedrock-runtime", region_name=BEDROCK_REGION)
        self.dynamodb = boto3.resource("dynamodb", region_name=BEDROCK_REGION)
        self.table = self.dynamodb.Table(DYNAMODB_TABLE)
        logger.info(f"✅ AWS 클라이언트 초기화 완료 (리전: {BEDROCK_REGION})")

    def read_excel(self, file_path: str) -> list[dict]:
        """Excel 파일에서 Q&A 데이터 읽기"""
        logger.info(f"📂 엑셀 파일 읽는 중: {file_path}")

        if not Path(file_path).exists():
            raise FileNotFoundError(f"❌ 파일을 찾을 수 없습니다: {file_path}")

        wb = openpyxl.load_workbook(file_path)
        sheet = wb[wb.sheetnames[SHEET_NAME]]

        rows = []
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is None or row[1] is None:
                continue

            rows.append({
                "id": str(uuid.uuid4()),  # 고유 ID 생성
                "question": str(row[0]).strip(),
                "answer": str(row[1]).strip(),
            })

        logger.info(f"✅ {len(rows)}개의 Q&A 쌍 읽음")
        return rows

    def embed_text(self, text: str) -> list[float]:
        """Bedrock Titan Embeddings으로 텍스트 임베딩"""
        try:
            response = self.bedrock.invoke_model(
                modelId=BEDROCK_MODEL_ID,
                contentType="application/json",
                accept="application/json",
                body=json.dumps({"inputText": text})
            )
            response_body = json.loads(response["body"].read())
            embedding = response_body["embedding"]
            logger.debug(f"  ✅ 임베딩 생성: {len(embedding)}차원")
            return embedding
        except Exception as e:
            logger.error(f"❌ 임베딩 생성 오류: {str(e)}")
            raise

    def upsert_to_dynamodb(self, qa_data: list[dict]) -> None:
        """DynamoDB에 데이터 저장"""
        logger.info(f"💾 DynamoDB에 {len(qa_data)}개 데이터 저장 중...")

        for i, row in enumerate(qa_data, start=1):
            try:
                # 임베딩 생성
                embedding = self.embed_text(row["question"])

                # DynamoDB에 저장
                item = {
                    "id": row["id"],
                    "question": row["question"],
                    "answer": row["answer"],
                    "embedding": embedding,  # List of floats
                }

                self.table.put_item(Item=item)
                logger.info(f"✅ [{i}/{len(qa_data)}] '{row['question'][:50]}...' 저장됨")

            except Exception as e:
                logger.error(f"❌ [{i}/{len(qa_data)}] 저장 실패: {str(e)}")
                raise

        logger.info("✅ 모든 데이터 저장 완료!")

    def run(self, file_path: Optional[str] = None) -> None:
        """전체 처리 흐름"""
        try:
            path = file_path or EXCEL_FILE
            qa_data = self.read_excel(path)
            self.upsert_to_dynamodb(qa_data)
            logger.info("🎉 임베딩 완료!")

        except Exception as e:
            logger.error(f"❌ 오류 발생: {str(e)}", exc_info=True)
            sys.exit(1)


if __name__ == "__main__":
    # 커맨드라인 인자로 파일 경로 받기
    excel_file = sys.argv[1] if len(sys.argv) > 1 else EXCEL_FILE

    ingestor = QAIngestor()
    ingestor.run(excel_file)
