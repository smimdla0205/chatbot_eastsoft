#!/usr/bin/env python3
"""
Q&A 데이터 임베딩 스크립트

용도:
1. Q&A.xlsx 파일 읽기
2. 각 질문을 임베딩 벡터로 변환
3. Supabase pgvector에 저장

실행:
python scripts/ingest.py
"""

import os
import sys
from pathlib import Path
from typing import Optional
import logging

try:
    import openpyxl
except ImportError:
    print("⚠️  openpyxl 필요. 설치 중...")
    os.system("pip install openpyxl")
    import openpyxl

try:
    from openai import OpenAI
except ImportError:
    print("⚠️  openai 필요. 설치 중...")
    os.system("pip install openai")
    from openai import OpenAI

try:
    import httpx
except ImportError:
    print("⚠️  httpx 필요. 설치 중...")
    os.system("pip install httpx")
    import httpx

from dotenv import load_dotenv

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# 환경 변수 로드
load_dotenv()

# 설정
EMBEDDING_MODEL = "text-embedding-3-small"
EXCEL_FILE = "data/Q&A.xlsx"  # 또는 사용자가 지정한 경로
SHEET_NAME = 0  # 첫 번째 시트


class QAIngestor:
    def __init__(self):
        """클라이언트 초기화"""
        self.openai = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))
        self.supabase_url = os.environ.get("SUPABASE_URL")
        self.supabase_key = os.environ.get("SUPABASE_ANON_KEY")

        if not all([self.openai.api_key, self.supabase_url, self.supabase_key]):
            raise ValueError("❌ 환경 변수 누락: OPENAI_API_KEY, SUPABASE_URL, SUPABASE_ANON_KEY")

    def read_excel(self, file_path: str) -> list[dict]:
        """Excel 파일에서 Q&A 데이터 읽기"""
        logger.info(f"📂 엑셀 파일 읽는 중: {file_path}")

        if not Path(file_path).exists():
            raise FileNotFoundError(f"❌ 파일을 찾을 수 없습니다: {file_path}")

        wb = openpyxl.load_workbook(file_path)
        sheet = wb[wb.sheetnames[SHEET_NAME]]

        rows = []
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # 열 이름 자동 감지 (Q&A 또는 Question/Answer 등)
            if row[0] is None or row[1] is None:
                continue

            rows.append({
                "id": i - 1,
                "question": str(row[0]).strip(),
                "answer": str(row[1]).strip(),
            })

        logger.info(f"✅ {len(rows)}개의 Q&A 쌍 읽음")
        return rows

    def embed_text(self, text: str) -> list[float]:
        """텍스트를 임베딩 벡터로 변환"""
        response = self.openai.embeddings.create(
            model=EMBEDDING_MODEL,
            input=text
        )
        return response.data[0].embedding

    def upsert_to_supabase(self, qa_data: list[dict]) -> None:
        """Supabase에 데이터 저장"""
        logger.info(f"💾 Supabase에 {len(qa_data)}개 데이터 저장 중...")

        url = f"{self.supabase_url}/rest/v1/qa_embeddings"
        headers = {
            "Authorization": f"Bearer {self.supabase_key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates",  # upsert
        }

        with httpx.Client() as client:
            for i, row in enumerate(qa_data):
                # 임베딩 생성
                embedding = self.embed_text(row["question"])

                payload = {
                    "id": row["id"],
                    "question": row["question"],
                    "answer": row["answer"],
                    "embedding": embedding,
                }

                try:
                    response = client.post(url, json=payload, headers=headers)
                    response.raise_for_status()
                    logger.info(f"✅ [{i+1}/{len(qa_data)}] '{row['question'][:50]}...' 저장됨")
                except Exception as e:
                    logger.error(f"❌ [{i+1}/{len(qa_data)}] 저장 실패: {str(e)}")
                    raise

        logger.info("✅ 모든 데이터 저장 완료!")

    def run(self, file_path: Optional[str] = None) -> None:
        """전체 처리 흐름"""
        try:
            path = file_path or EXCEL_FILE
            qa_data = self.read_excel(path)
            self.upsert_to_supabase(qa_data)
            logger.info("🎉 임베딩 완료!")

        except Exception as e:
            logger.error(f"❌ 오류 발생: {str(e)}", exc_info=True)
            sys.exit(1)


if __name__ == "__main__":
    # 커맨드라인 인자로 파일 경로 받기
    excel_file = sys.argv[1] if len(sys.argv) > 1 else EXCEL_FILE

    ingestor = QAIngestor()
    ingestor.run(excel_file)
